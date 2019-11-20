# -*- coding: utf-8 -*-


# Clase 2

# librerias a utilizar: pandas, numpy
# pandas: para tratiemto de data y archivos
# numpy : para todo lo relacionado a cálculo matemático y vectores matemáticos.

import pandas as pd
import numpy as np

# leer archivos:
data = pd.read_excel("C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx")
# esta variable pasa a ser un Data Frame
type(data)
data
print(data)
# veamos las columnas
data.columns

# veamos las filas
data.index

# podemos cambiar el indice
data = pd.read_excel("C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx", index_col = "Nombre")

# pandas nos ofrece una función para obtener un resumen y análisis rapido de la data que tenemos:
data.describe()

# podemos crear data frames a partir de diccionarios:
data_dict

# calculemos el indice de masa corporal (IMC):
def IMC(peso,altura):
    return (altura**2)/peso

# agreguemos la data al data frame creado:
data['IMC'] = IMC(data['peso'],data['altura'])
data
# guardemos este archivo en un excel:
data.to_excel("C:\Program BI\Python ver 2.0\Clase 2\IMC.xlsx")

# Para guardar en el mismo Excel
import openpyxl
doc = openpyxl.load_workbook('C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx')
hoja = doc['Datos']
hoja.max_column
hoja.max_row
# agreguemos un nuevo registro y luego guardamos y cerramos:
hoja.append(['Pedro',24,100,1.9])
doc.save("C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx")
doc.close


# ¿Como podríamos guardar el IMC en el mismo excel?:
doc = openpyxl.load_workbook('C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx')
hoja = doc['Datos']
hoja.max_column
hoja.max_row
data = pd.read_excel("C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx", index_col = "Nombre")
data['IMC'] = IMC(data['peso'],data['altura'])
hoja.cell(1,4).value = 'IMC'
IMC = list(data['IMC'])
for i in range(2,5):
    hoja.cell(i,4).value = IMC[i-2]

doc.save("C:\Program BI\Python ver 2.0\Clase 2\personas.xlsx")
doc.close


# Desafío, utilizando el archivo productos.xlsx, y dado un diccionario, agregar la 
# columna comprador, que será la persona que compre aquel artículo.

compradores = {'mesa': 'Luis', 'libro' : 'pedro', 'polera': 'juan','zapatos':'luis'}

print(compradores)



# 




