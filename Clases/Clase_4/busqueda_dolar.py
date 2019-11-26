# -*- coding: utf-8 -*-

# Para que busque la ruta sola del archivo:
import os
#os.chdir('C:\Program BI\Python ver 2.0')
basedir = __file__[:__file__.rfind('/')+1]
if basedir != '': os.chdir(basedir)

# Importamos libreria que nos ayuda a conectarnos a la web
from bs4 import BeautifulSoup as bs
import requests
import datetime as dt
import openpyxl

# Obtenemos el Dolar del Banco Central
def trae_dolar():
    main_url = "https://www.bcentral.cl"
    req = requests.get(main_url)
    soup = bs(req.text, "html.parser")
    title = soup.find("div", class_ = "tab-pane active")    
    #len(lista)
    x = str(list(title)[1])  
    frase = 'Dólar Observado</td>\n<td>'
    pos_1 = x.find(frase)
    pos_2 = x[pos_1:].find('</td>\n</tr>')+pos_1
    new_x = x[pos_1:pos_2]
    pos_3 = new_x.find('<td>')
    valor = float(new_x[pos_3+4:].replace(',','.'))
    return (valor)


# Abrimos el libro y en la hoja respectiva
def guada_dolar():
    dolar = trae_dolar()
    doc = openpyxl.load_workbook('Dolar.xlsx')
    hoja = doc['valor_dolar']
    #hoja.max_column
    #hoja.max_row
    fecha_hoy = dt.datetime.today()
    fecha_hoy.year    
    # poner controles sobre la carga de informacion:    
    condicion = True
    for i in range(1,hoja.max_row+1):
        fecha_aux = dt.datetime(fecha_hoy.year,fecha_hoy.month,fecha_hoy.day)
        if hoja['A'+str(i)].value == fecha_aux:
            condicion = False
            break            
    if condicion:
        hoja.append([fecha_hoy,valor])
    else:
        print('Ya cargaste el dolar!!')
    # Por si quisiera grabar el dolar en una posición fija
    # hoja['A3'].value = valor
    doc.save("Dolar.xlsx")
    doc.close
# Desafio: intentar sacar el valor del Euro desde el banco central











