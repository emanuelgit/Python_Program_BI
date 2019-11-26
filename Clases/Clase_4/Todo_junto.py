# -*- coding: utf-8 -*-
"""
Created on Mon Nov 25 13:49:13 2019

@author: Emanuel&Mara
"""
from win32com import client
#import pandas as pd
import datetime as dt
import os
from bs4 import BeautifulSoup as bs
import requests
import openpyxl
#os.chdir('C:\Program BI\Python ver 2.0')
basedir = __file__[:__file__.rfind('/')+1]
if basedir != '': os.chdir(basedir)

def trae_dolar():
    main_url = "https://www.bcentral.cl"
    req = requests.get(main_url)
    soup = bs(req.text, "html.parser")
    title = soup.find("div", class_ = "tab-pane active")    
    #len(lista)
    x = str(list(title)[1])  
    frase = 'Dólar Observado</td>\n<td>'
    pos_1 = x.find(frase)
    pos_2 = x[pos_1:].find('</td>\n</tr>')+pos_1
    new_x = x[pos_1:pos_2]
    pos_3 = new_x.find('<td>')
    valor = float(new_x[pos_3+4:].replace(',','.'))
    return (valor)


# Abrimos el libro y en la hoja respectiva
def guada_dolar():
    valor = trae_dolar()
    doc = openpyxl.load_workbook('Dolar.xlsx')
    hoja = doc['valor_dolar']
    #hoja.max_column
    #hoja.max_row
    fecha_hoy = dt.date.today()
    fecha_hoy.year    
    # poner controles sobre la carga de informacion:    
    condicion = True
    for i in range(1,hoja.max_row+1):
        fecha_aux = dt.datetime(fecha_hoy.year,fecha_hoy.month,fecha_hoy.day)
        if hoja['A'+str(i)].value == fecha_aux:
            condicion = False
            break            
    if condicion:
        hoja.append([fecha_hoy,valor])
    else:
        print('Ya cargaste el dolar!!')
    # Por si quisiera grabar el dolar en una posición fija
    # hoja['A3'].value = valor
    doc.save("Dolar.xlsx")
    doc.close
    return(valor)



def enviar_mail(subject, body, mails, attachment_paths=None):    
    olMailItem = 0x0
    obj = client.gencache.EnsureDispatch("Outlook.Application")
    newMail = obj.CreateItem(olMailItem)
    newMail.Subject = subject    
    cadena = ""
    for mail in mails:
        mail = mail+";"
        cadena = cadena+mail
    cadena=cadena[:-1] # los mails se mandan separados por ; en una cadena
    newMail.To = cadena
    #newMail.CC = cadena
    #newMail.CCO = cadena
    newMail.BCC = 'moliva@programbi.cl'
    newMail.Body = body
    if attachment_paths:
        for att in attachment_paths:
            newMail.Attachments.Add(Source=att)
    try:
        newMail.Send()
        print("Correo enviado")
    except:
        print("Fallo en el envío del correo")
        


#enviar_mail('El dolar del día','El dolar fue:',
#                ['eberrocal@programbi.cl'],
#                ['C:/Program BI/Python ver 2.0/Clase 4/Dolar.xlsx'])  

        
        
#if __name__ == '__main__':
#    guada_dolar()
#    enviar_mail('El dolar del día','El dolar fue:',
#                ['emanuel.berrocal@gmail.com'],
#                ['C:/Program BI/Python ver 2.0/Clase 4/Dolar.xlsx'])  
    
    
    
    
    
    
    