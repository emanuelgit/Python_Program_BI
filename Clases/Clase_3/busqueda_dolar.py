# -*- coding: utf-8 -*-

# Para que busque la ruta sola del archivo:
import os
basedir = __file__[:__file__.rfind('/')+1]
if basedir != '': os.chdir(basedir)

# Importamos libreria que nos ayuda a conectarnos a la web
from bs4 import BeautifulSoup as bs
import requests
import datetime as dt

# Obtenemos el Dolar del Banco Central
main_url = "https://www.bcentral.cl"
req = requests.get(main_url)
soup = bs(req.text, "html.parser")
title = soup.find("div", class_ = "tab-pane active")
lista = list(title)
#len(lista)
x = str(list(title)[1])  
frase = 'Dólar Observado</td>\n<td>'
pos_1 = x.find(frase)
pos_2 = x[pos_1:].find('</td>\n</tr>')+pos_1
new_x = x[pos_1:pos_2]
pos_3 = new_x.find('<td>')
valor = float(new_x[pos_3+4:].replace(',','.'))


# Abrimos el libro y en la hoja respectiva
import openpyxl
doc = openpyxl.load_workbook('Dolar.xlsx')
hoja = doc['valor_dolar']
#hoja.max_column
#hoja.max_row
fecha_hoy = dt.date.today()
hoja.append([fecha_hoy,valor])

doc.save("Dolar.xlsx")
doc.close

# Desafio: intentar sacar el valor del Euro desde el banco central











