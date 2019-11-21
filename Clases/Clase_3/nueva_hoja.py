# -*- coding: utf-8 -*-

import openpyxl
# Empezamos creando un libro vacío, que ya viene con una hoja
#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('Dolar.xlsx')

# Creo otra hoja en el libro
wb.create_sheet(index=1, title='Otra hoja')
# Y así se borra, si hace falta
#wb.remove_sheet(wb.get_sheet_by_name('Otra hoja'))

# Estilo a las celdas
# from openpyxl.styles import Font


# Guardo la hoja
wb.save('Dolar.xlsx')
wb.close

# En la nueva hoja agregar el Euro con su fecha, respectiva